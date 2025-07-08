import os
import math
from pyInvoice.data_manager import *
from tqdm import tqdm
from openpyxl import load_workbook, Workbook
from pyInvoice import constant
from pyInvoice import utils
from openpyxl.worksheet.pagebreak import Break
from copy import copy
from datetime import datetime
from openpyxl.utils.datetime import from_excel

import win32com.client
from pathlib import Path


class Cell:
    def __init__(self, row, column, value=""):
        self.row = row
        self.column = column
        self.value = value
        
class Template:
    def __init__(
        self, 
        name: str, 
        label: str,
        # sheet_name: str,
        base_file_path: str, 
        output_dir: str,
        setting_dict,
        date_cell: Cell = None, 
        nouhin_no_cell: Cell = None, 
        shop_name_cell: Cell = None,         
        brand_name_cell: Cell = None,
        packing_no_cell: Cell = None, 
        start_row_item: int = 13,
        max_row_item: int = 25,
        print_area: tuple = (43, 9),
        merge_cells: list = None,
    ):
        self.date_cell       = date_cell or Cell(2, 8)
        self.nouhin_no_cell  = nouhin_no_cell or Cell(3, 8)
        self.shop_name_cell  = shop_name_cell or Cell(7, 3, setting_dict.get(constant.attr_str_shop_key))
        self.brand_name_cell = brand_name_cell or Cell(8, 3, setting_dict.get(constant.attr_str_brand_key))
        self.packing_no_cell = packing_no_cell or Cell(7, 5)
        self.start_row_item  = start_row_item
        self.max_row_item    = max_row_item
        self.print_area      = print_area
        self.merge_cells     = merge_cells or [
            [Cell(1, 1), Cell(1, 9)],
            [Cell(2, 6), Cell(2, 7)],
            [Cell(2, 8), Cell(2, 9)],
            [Cell(3, 6), Cell(3, 7)],
            [Cell(3, 2), Cell(3, 4)],
            [Cell(3, 8), Cell(3, 9)],
            [Cell(7, 3), Cell(7, 4)],
            [Cell(7, 5), Cell(8, 5)],
            [Cell(8, 3), Cell(8, 4)],
        ]
        self.name           = name
        self.label          = label
        # self.sheet_name     = sheet_name
        self.base_file_path = base_file_path
        self.output_dir     = output_dir
 
    def write(self, packing_data : PackingList, item_master : ItemMaster, logger):        
        #ps = packing_data.packing_set  
        ps = dict(sorted(packing_data.packing_set.items()))

        error_items = list()

        # 新しいファイルとして保存
        
        new_file_path = os.path.join(self.output_dir, f"output.xlsx")
        relative_path = Path(new_file_path)
        absolute_path = relative_path.resolve()


        wb = load_workbook(self.base_file_path)
        first_sheet_name = wb.sheetnames[0]  # 先頭のシート名
        ws = wb[first_sheet_name]     
        
        # コピー元範囲
        nrow, ncol = self.print_area[0], self.print_area[1]
        start_row = 1
        n_page = sum([math.ceil(len(items)/self.max_row_item) for k, items in ps.items()])
        for i in range(n_page-1):
            for row in range(start_row, start_row + nrow):  # A1〜H43
                for col in range(1, ncol+1):  # A〜H
                    source_cell = ws.cell(row=row, column=col)
                    target_row = row + nrow
                    target_cell = ws.cell(row=target_row, column=col)

                    # 値とフォント・書式をコピー
                    target_cell.value = source_cell.value
                    if source_cell.has_style:
                        target_cell.font = copy(source_cell.font)
                        target_cell.border = copy(source_cell.border)
                        target_cell.fill = copy(source_cell.fill)
                        target_cell.number_format = source_cell.number_format
                        target_cell.protection = copy(source_cell.protection)
                        target_cell.alignment = copy(source_cell.alignment)
            start_row += nrow

        #請求書作成
        start_row = 1
        for i, (packing_no, packing_items) in tqdm(enumerate(ps.items()), total=len(ps), desc="【請求書作成中】"):
            many_page = len(packing_items) > self.max_row_item            
            for k in range(math.ceil(len(packing_items)/self.max_row_item)):

                ws.cell(row=self.date_cell.row + start_row - 1, column=self.date_cell.column).value = utils.to_datetime_from_excel(packing_data.date)
                ws.cell(row=self.date_cell.row + start_row - 1, column=self.date_cell.column).number_format = 'yyyy/mm/dd'                
                ws.cell(row=self.nouhin_no_cell.row + start_row - 1, column=self.nouhin_no_cell.column).value = (
                    f"{packing_data.nouhin_no}{str(packing_no).zfill(3)}_{k+1}{self.label}"
                    if many_page else f"{packing_data.nouhin_no}{str(packing_no).zfill(3)}{self.label}"
                )
                ws.cell(row=self.packing_no_cell.row + start_row - 1, column=self.packing_no_cell.column).value =  f"箱NO【{packing_no}】"
                ws.cell(row=self.brand_name_cell.row + start_row - 1, column=self.brand_name_cell.column).value = self.brand_name_cell.value
                ws.cell(row=self.shop_name_cell.row + start_row - 1, column=self.shop_name_cell.column).value = self.shop_name_cell.value


                for j, packing in enumerate(packing_items[(k*self.max_row_item):(k+1)*self.max_row_item]):
                    row = self.start_row_item+j+start_row-1
                    packing: Packing
                    jan = packing.jan
                    item: Item = item_master.item_dict.get(jan, Item())
                    
                    if item:  # Noneチェック
                        ws.cell(row=row, column=2).value = item.brand_code
                        ws.cell(row=row, column=3).value = item.color
                        ws.cell(row=row, column=4).value = item.size
                        ws.cell(row=row, column=5).value = item.name
                        ws.cell(row=row, column=6).value = item.jan
                        ws.cell(row=row, column=7).value = packing.amount
                        ws.cell(row=row, column=8).value = item.price
                        ws.cell(row=row, column=9).value = item.price * packing.amount
                        if(not item.brand_code or not item.color or not item.size):
                            error_items.append(item)
                
                if(start_row > 1):
                    for merge_cell in self.merge_cells:
                        p0, p1 = merge_cell[0], merge_cell[1]
                        ws.merge_cells(
                            start_row=start_row-1+p0.row, 
                            start_column=p0.column, 
                            end_row=start_row-1+p1.row, 
                            end_column=p1.column
                        )

                if(i + 1 != len(ps)):
                    ws.row_breaks.append(Break(id=nrow * (i + 1)))
                start_row += nrow


        ws.print_area = f"A1:I{start_row}"            
        wb.save(absolute_path)
        self.convert_excel_to_pdf_win(absolute_path)

        if(len(error_items) > 0):
            # 新規ブック作成
            wb = Workbook()
            ws = wb.active
            ws.title = "エラーリスト"  # 任意のシート名に変更可能

            # ヘッダー行を追加
            ws.append(["JANコード", "カラーの設定", "ブランド品番の設定", "サイズの設定"])
            for item in error_items:
                ws.append([item.jan, utils.judge(item.color) , utils.judge(item.brand_code), utils.judge(item.size)])

            error_list_file_path = os.path.join(self.output_dir, f"error_list.xlsx")
            wb.save(error_list_file_path)
            
            logger.warning(f"設定に不備があるマスタが存在します.{error_list_file_path}を確認してください")
        else:
            logger.info(f"正常終了")
        return 
    
    def convert_excel_to_pdf_win(excel_path, pdf_path=None):
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        wb = excel.Workbooks.Open(excel_path)

        if not pdf_path:
            pdf_path = excel_path.replace(".xlsx", ".pdf")

        wb.ExportAsFixedFormat(0, pdf_path)
        wb.Close(False)
        excel.Quit()

        print(f"✅ PDF変換成功: {pdf_path}")
        return